import openpyxl
def get_rows(file,sheet_name):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    return (sheet.max_row)
def get_columns(file,sheet_name):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    return (sheet.max_column)
def read_data(file,sheet_name,row_num,column_num):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_num,column=column_num).value
def write_data(file,sheet_name,row_num,column_num,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num,column=column_num).value=data
    workbook.save(file)